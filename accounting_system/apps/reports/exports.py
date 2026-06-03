"""Excel (openpyxl) export builders."""
from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="4F46E5")
HEADER_FONT = Font(color="FFFFFF", bold=True)
MONEY_FMT = "#,##0.00"


def _autofit(ws):
    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(length + 4, 48)


def _style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _to_bytes(wb) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def transactions_xlsx(transactions) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Транзакции"
    headers = ["Дата", "Тип", "Сумма", "Валюта", "Категория", "Клиент", "Описание"]
    ws.append(headers)
    for tx in transactions:
        ws.append([
            tx.date.strftime("%d.%m.%Y"),
            tx.get_type_display(),
            float(tx.amount),
            tx.currency,
            tx.category.name if tx.category_id else "",
            tx.client.name if tx.client_id else "",
            tx.description,
        ])
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=3).number_format = MONEY_FMT
    _style_header(ws, len(headers))
    _autofit(ws)
    return _to_bytes(wb)


def pnl_xlsx(rows) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "P&L"
    headers = ["Период", "Выручка", "Расходы", "Прибыль", "Маржа %"]
    ws.append(headers)
    for r in rows:
        ws.append([r["label"], float(r["revenue"]), float(r["expense"]),
                   float(r["profit"]), round(float(r["margin"]), 1)])
    for row in range(2, ws.max_row + 1):
        for col in (2, 3, 4):
            ws.cell(row=row, column=col).number_format = MONEY_FMT
    _style_header(ws, len(headers))
    _autofit(ws)
    return _to_bytes(wb)


def invoices_xlsx(invoices) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Счета"
    headers = ["Номер", "Клиент", "Статус", "Дата", "Срок", "Подытог", "НДС", "Итого", "Валюта"]
    ws.append(headers)
    for inv in invoices:
        ws.append([
            inv.number, inv.client.name, inv.get_status_display(),
            inv.issue_date.strftime("%d.%m.%Y"), inv.due_date.strftime("%d.%m.%Y"),
            float(inv.subtotal), float(inv.tax_amount), float(inv.total), inv.currency,
        ])
    for row in range(2, ws.max_row + 1):
        for col in (6, 7, 8):
            ws.cell(row=row, column=col).number_format = MONEY_FMT
    _style_header(ws, len(headers))
    _autofit(ws)
    return _to_bytes(wb)
