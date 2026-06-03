"""Parse shop order exports (CSV / Excel) — only accounting-relevant columns."""
from __future__ import annotations

import csv
import io
import re
from datetime import datetime
from decimal import Decimal, InvalidOperation

from django.utils import timezone

from .order_data import PAID_STATUSES, ShopOrderData

REQUIRED_FOR_IMPORT = frozenset({"id", "status", "total_price", "created_at"})

COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    "id": ("id", "order_id", "№", "номер", "номер_заказа"),
    "email": ("email", "e_mail", "e-mail", "почта"),
    "first_name": ("first_name", "firstname", "имя", "first"),
    "last_name": ("last_name", "lastname", "фамилия", "last"),
    "phone": ("phone", "tel", "telephone", "телефон", "тел"),
    "status": ("status", "статус", "state"),
    "total_price": ("total_price", "total", "sum", "amount", "сумма", "итого", "totalprice"),
    "gift_card_debit": ("gift_card_debit", "gift_card", "сертификат", "giftcard_debit"),
    "payable_amount": ("payable_amount", "payable", "к_оплате", "payableamount"),
    "created_at": (
        "created_at", "created", "date", "datetime", "дата", "дата_создания", "createdat",
    ),
}

FIXED_COLUMN_ORDER = (
    "id", "email", "first_name", "last_name", "phone", "country",
    "address_line1", "address_line2", "city", "region", "postal_code",
    "delivery_method", "order_note", "total_price", "created_at", "user_id",
    "status", "tracking_number", "gift_card_debit", "payable_amount",
)


def read_file_bytes(file_obj) -> bytes:
    """Read upload or path handle reliably (Django may leave cursor at EOF)."""
    if isinstance(file_obj, (bytes, bytearray)):
        return bytes(file_obj)
    if hasattr(file_obj, "seek"):
        try:
            file_obj.seek(0)
        except (OSError, ValueError):
            pass
    data = file_obj.read()
    if not data and hasattr(file_obj, "chunks"):
        data = b"".join(chunk for chunk in file_obj.chunks() if chunk)
    return data or b""


def _guess_filename(raw: bytes, filename: str) -> str:
    name = (filename or "").lower().strip()
    if name.endswith((".csv", ".tsv", ".txt", ".xlsx")):
        return name
    if raw.startswith(b"PK\x03\x04"):
        return "upload.xlsx"
    return "upload.csv"


def _clean_header(key: str) -> str:
    k = (key or "").strip().strip("\ufeff").strip("\r\n")
    k = k.strip('"\'`')
    k = re.sub(r"[^\w\u0400-\u04ff]+", "_", k, flags=re.UNICODE)
    k = re.sub(r"_+", "_", k).strip("_").lower()
    return k


def _canonical_name(clean_key: str) -> str | None:
    if not clean_key:
        return None
    for canon, aliases in COLUMN_ALIASES.items():
        if clean_key == canon or clean_key in aliases:
            return canon
    if clean_key.endswith("_id") and clean_key.rstrip("_id") in ("", "order", "products_order"):
        return "id"
    return None


def _map_row(raw: dict) -> dict[str, str]:
    out: dict[str, str] = {}
    for raw_key, value in raw.items():
        canon = _canonical_name(_clean_header(str(raw_key)))
        if not canon:
            continue
        if value is None:
            text = ""
        elif isinstance(value, datetime):
            text = value.strftime("%Y-%m-%d %H:%M:%S")
        else:
            text = str(value).strip()
        if text.startswith('"') and text.endswith('"'):
            text = text[1:-1].strip()
        out[canon] = text
    return out


def _header_score(keys: list[str]) -> int:
    mapped = {_canonical_name(_clean_header(k)) for k in keys}
    mapped.discard(None)
    return len(REQUIRED_FOR_IMPORT & mapped)


def _parse_decimal(value: str) -> Decimal:
    cleaned = (value or "").replace("\u00a0", "").replace(" ", "")
    if cleaned.count(",") == 1 and cleaned.count(".") == 0:
        cleaned = cleaned.replace(",", ".")
    elif "," in cleaned and "." in cleaned:
        cleaned = cleaned.replace(",", "")
    if not cleaned:
        return Decimal("0")
    return Decimal(cleaned)


def _parse_datetime(value: str) -> datetime:
    value = (value or "").strip()
    if not value:
        raise ValueError("пустая дата")
    for fmt in (
        "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M:%S.%f",
        "%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M:%S.%f",
        "%Y-%m-%d", "%d.%m.%Y %H:%M:%S", "%d.%m.%Y",
        "%d/%m/%Y %H:%M:%S", "%d/%m/%Y",
        "%m/%d/%Y %H:%M:%S", "%m/%d/%Y",
    ):
        try:
            dt = datetime.strptime(value[:26], fmt)
            if timezone.is_naive(dt):
                return timezone.make_aware(dt, timezone.get_current_timezone())
            return dt
        except ValueError:
            continue
    try:
        dt = datetime.fromisoformat(value.replace("Z", "+00:00"))
        if timezone.is_naive(dt):
            return timezone.make_aware(dt, timezone.get_current_timezone())
        return dt
    except ValueError as exc:
        raise ValueError(f"некорректная дата «{value}»") from exc


def row_to_order(data: dict[str, str], line_no: int) -> ShopOrderData:
    missing = REQUIRED_FOR_IMPORT - set(data.keys())
    if missing:
        found = ", ".join(sorted(data.keys())) or "—"
        raise ValueError(
            f"нет колонок: {', '.join(sorted(missing))} (в строке есть: {found})"
        )

    status = data["status"].lower()
    if status not in PAID_STATUSES:
        raise ValueError(f"статус «{status}» не импортируется (только оплаченные)")

    amount = _parse_decimal(data.get("total_price", "0"))
    if amount <= 0:
        amount = _parse_decimal(data.get("payable_amount", "0")) + _parse_decimal(
            data.get("gift_card_debit", "0")
        )
    if amount <= 0:
        raise ValueError("сумма заказа должна быть > 0")

    order_id = int(str(data["id"]).split(".")[0])

    return ShopOrderData(
        id=order_id,
        email=data.get("email") or "",
        first_name=data.get("first_name") or "",
        last_name=data.get("last_name") or "",
        phone=data.get("phone") or "",
        status=status,
        total_price=_parse_decimal(data.get("total_price") or "0") or amount,
        gift_card_debit=_parse_decimal(data.get("gift_card_debit") or "0"),
        payable_amount=_parse_decimal(data.get("payable_amount") or "0"),
        created_at=_parse_datetime(data["created_at"]),
    )


def _decode_text(raw: bytes) -> str:
    if not raw:
        return ""
    if raw.startswith(b"\xff\xfe") or raw.startswith(b"\xfe\xff"):
        return raw.decode("utf-16")
    # Excel on Windows often saves CSV as UTF-16 LE without BOM
    if len(raw) > 4 and raw[1:2] == b"\x00" and raw[3:4] == b"\x00":
        try:
            return raw.decode("utf-16-le")
        except UnicodeDecodeError:
            pass
    for encoding in ("utf-8-sig", "utf-8", "cp1251", "latin-1"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace")


def _detect_csv_layout(lines: list[str]) -> tuple[str, int, list[str]]:
    best = (";", 0, [], -1)
    sample_lines = lines[: min(15, len(lines))]

    for delim in (";", ",", "\t", "|"):
        for idx, line in enumerate(sample_lines):
            if not line.strip():
                continue
            try:
                row = next(csv.reader([line], delimiter=delim))
            except csv.Error:
                continue
            score = _header_score([str(c) for c in row])
            if score > best[3]:
                best = (delim, idx, [str(c) for c in row], score)

    delim, idx, keys, score = best
    if score < len(REQUIRED_FOR_IMPORT):
        for line in sample_lines:
            for d in (";", ",", "\t"):
                try:
                    row = next(csv.reader([line], delimiter=d))
                except csv.Error:
                    continue
                if len(row) >= len(FIXED_COLUMN_ORDER) and str(row[0]).strip().isdigit():
                    return d, -1, list(FIXED_COLUMN_ORDER)

        preview = sample_lines[0][:120] if sample_lines else ""
        raise ValueError(
            "Не удалось распознать заголовки. В phpMyAdmin: Экспорт → CSV, "
            f"колонки id, status, total_price, created_at. Первая строка: «{preview}…»"
        )
    return delim, idx, keys


def _read_csv_rows(raw: bytes) -> list[dict]:
    text = _decode_text(raw)
    lines = text.splitlines()
    if not lines:
        raise ValueError("Файл пуст.")

    delim, header_idx, fieldnames = _detect_csv_layout(lines)
    data_lines = lines if header_idx < 0 else lines[header_idx + 1 :]

    out: list[dict] = []
    for row in csv.reader(data_lines, delimiter=delim):
        if not any(str(c).strip() for c in row):
            continue
        padded = list(row) + [""] * max(0, len(fieldnames) - len(row))
        raw_row = dict(zip(fieldnames, padded[: len(fieldnames)]))
        mapped = _map_row(raw_row)
        if mapped:
            out.append(mapped)
    return out


def _read_xlsx_rows(file_obj) -> list[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(file_obj, read_only=True, data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not all_rows:
        raise ValueError("Excel-файл пуст.")

    header_idx, best_score = 0, -1
    for idx, row in enumerate(all_rows[:15]):
        score = _header_score([str(c or "") for c in row])
        if score > best_score:
            best_score, header_idx = score, idx

    if best_score < len(REQUIRED_FOR_IMPORT):
        if all_rows[0] and str(all_rows[0][0] or "").strip().isdigit():
            header_idx = -1
        else:
            raise ValueError(
                "Не удалось распознать заголовки в Excel. "
                "Нужны колонки: id, status, total_price, created_at."
            )

    if header_idx < 0:
        fieldnames = list(FIXED_COLUMN_ORDER)
        data_rows = all_rows
    else:
        fieldnames = [str(c or f"col_{i}") for i, c in enumerate(all_rows[header_idx])]
        data_rows = all_rows[header_idx + 1 :]

    out: list[dict] = []
    for row in data_rows:
        if not any(c is not None and str(c).strip() for c in row):
            continue
        padded = list(row) + [None] * max(0, len(fieldnames) - len(row))
        raw_row = {fieldnames[i]: padded[i] for i in range(len(fieldnames))}
        mapped = _map_row(raw_row)
        if mapped:
            out.append(mapped)
    return out


def parse_orders_file(file_obj, filename: str) -> tuple[list[ShopOrderData], list[str]]:
    raw = read_file_bytes(file_obj)
    if not raw:
        raise ValueError(
            "Файл пуст или не загрузился. Выберите CSV снова и нажмите «Импортировать»."
        )
    name = _guess_filename(raw, filename)
    if name.endswith(".xlsx"):
        raw_rows = _read_xlsx_rows(io.BytesIO(raw))
    elif name.endswith((".csv", ".txt", ".tsv")):
        raw_rows = _read_csv_rows(raw)
    else:
        raise ValueError("Поддерживаются файлы .csv, .tsv и .xlsx")

    orders: list[ShopOrderData] = []
    errors: list[str] = []
    for i, row in enumerate(raw_rows):
        line_no = i + 2
        try:
            orders.append(row_to_order(row, line_no))
        except (ValueError, InvalidOperation, KeyError) as exc:
            errors.append(f"Строка {line_no}: {exc}")
    return orders, errors
